import csv
import io
from django.http import HttpResponse
from openpyxl import Workbook
from rest_framework import generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from .models import Product, ProductVariant
from .serializers import ProductSerializer, ProductVariantSerializer
from .utils import parse_and_import_csv, parse_and_import_accessories_csv
from brands.models import BrandLogo
from categories.models import USPCategory


class ProductListView(generics.ListAPIView):
    queryset = Product.objects.prefetch_related('variants').all()
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.query_params.get('search', '')
        brand = self.request.query_params.get('brand', '')
        product_type = self.request.query_params.get('product_type', '')
        if search:
            qs = qs.filter(commercial_name__icontains=search) | qs.filter(brand__icontains=search) | qs.filter(product_id__icontains=search)
        if brand:
            qs = qs.filter(brand__iexact=brand)
        if product_type:
            qs = qs.filter(product_type=product_type)
        return qs


class ProductDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Product.objects.prefetch_related('variants').all()
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]


class ProductBatchDeleteView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'detail': 'ids wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)
        deleted, _ = Product.objects.filter(id__in=ids).delete()
        return Response({'deleted': deleted})


class ProductVariantDetailView(generics.RetrieveUpdateAPIView):
    queryset = ProductVariant.objects.all()
    serializer_class = ProductVariantSerializer
    permission_classes = [IsAuthenticated]


class ImportProductCSVView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)
        if not file.name.endswith('.csv'):
            return Response({'detail': 'File must be a CSV.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            result = parse_and_import_csv(file)
            return Response(result, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class ImportAccessoriesCSVView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)
        if not file.name.endswith('.csv'):
            return Response({'detail': 'File must be a CSV.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            result = parse_and_import_accessories_csv(file)
            return Response(result, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)


def resolve_usp_icon(usp_text, categories):
    """Match USP text against category keywords, return icon_url or empty string."""
    if not usp_text:
        return ''
    text_lower = usp_text.lower()
    for cat in categories:
        for kw in cat.keywords:
            if kw.lower() in text_lower:
                return cat.icon_url
    return ''


def fmt_installment(value):
    """Format installment integer as 'Rp 123.456/bulan'"""
    if not value:
        return ''
    formatted = f"{int(value):,}".replace(',', '.')
    return f"Rp {formatted}/bulan"


def fmt_price(value):
    """Format price: remove trailing 000 (divide by 1000)"""
    if not value:
        return ''
    return str(int(value) // 1000)


class ExportDemoItemCSVView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Pre-load brand logos into dict {brand_lower: logo_url}
        brand_logos = {b.brand.lower(): b.logo_url for b in BrandLogo.objects.all()}

        # Pre-load USP categories (exclude Blibli White/Black which have no keywords role)
        categories = list(USPCategory.objects.exclude(keywords=[]))

        products = Product.objects.filter(
            product_type=Product.PRODUCT_TYPE_DEMO
        ).prefetch_related('variants').order_by('product_id')

        # Apply optional filters
        ids_param = request.query_params.get('ids', '')
        search = request.query_params.get('search', '')
        brand = request.query_params.get('brand', '')
        if ids_param:
            id_list = [i.strip() for i in ids_param.split(',') if i.strip()]
            products = products.filter(id__in=id_list)
        else:
            if search:
                products = products.filter(commercial_name__icontains=search) | \
                           products.filter(brand__icontains=search)
            if brand:
                products = products.filter(brand__iexact=brand)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Demo Item'

        headers = [
            'Barcode', 'Product name', 'Specification', 'Price', 'Original Price',
            'Image Url', 'Discount Price',
            'Customize1', 'Customize2', 'Customize3', 'Customize4',
            'Customize5', 'Customize6', 'Customize7', 'Customize8',
            'Customize9', 'Customize10', 'Customize11', 'Customize12', 'Customize13',
            'Customize14', 'Customize15', 'Customize16', 'Customize17', 'Customize18',
            'Customize19', 'Customize20',
            'Customize21', 'Customize22', 'Customize23', 'Customize24', 'Customize25',
        ]
        ws.append(headers)

        def fmt_storage(val):
            if not val:
                return ''
            try:
                n = int(val)
                return f"{n // 1000} TB" if n >= 1000 else f"{n} GB"
            except (ValueError, TypeError):
                return val

        def ram_rom(v):
            if not v:
                return ''
            parts = []
            if v.ram:
                parts.append(fmt_storage(v.ram))
            if v.rom:
                parts.append(fmt_storage(v.rom))
            return ' / '.join(parts) if parts else ''

        for p in products:
            variants = sorted(p.variants.all(), key=lambda v: v.variant_number)
            while len(variants) < 5:
                variants.append(None)

            def var(i, _v=variants):
                return _v[i] if i < len(_v) else None

            image_url = brand_logos.get(p.brand.lower(), '')
            usps = [p.usp_1, p.usp_2, p.usp_3, p.usp_4]
            usp_icons = [resolve_usp_icon(u, categories) for u in usps]

            ws.append([
                p.product_id,
                p.commercial_name,
                p.colour,
                fmt_price(var(0).unit_price if var(0) else None),
                fmt_price(var(1).unit_price if var(1) else None),
                image_url,
                fmt_price(var(2).unit_price if var(2) else None),
                # Customize 1-4: USP icons
                usp_icons[0], usp_icons[1], usp_icons[2], usp_icons[3],
                # Customize 5-8: USP text
                p.usp_1, p.usp_2, p.usp_3, p.usp_4,
                # Customize 9-13: RAM/ROM var 1-5
                ram_rom(var(0)), ram_rom(var(1)), ram_rom(var(2)), ram_rom(var(3)), ram_rom(var(4)),
                # Customize 14-18: installment var 1-5
                fmt_installment(var(0).installment if var(0) else None),
                fmt_installment(var(1).installment if var(1) else None),
                fmt_installment(var(2).installment if var(2) else None),
                fmt_installment(var(3).installment if var(3) else None),
                fmt_installment(var(4).installment if var(4) else None),
                # Customize 19-20: price var 4-5
                fmt_price(var(3).unit_price if var(3) else None),
                fmt_price(var(4).unit_price if var(4) else None),
                # Customize 21-25: empty
                '', '', '', '', '',
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="esl_demo_export.xlsx"'
        return response


class ExportAccessoriesItemView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        brand_logos = {b.brand.lower(): b.logo_url for b in BrandLogo.objects.all()}

        products = Product.objects.filter(
            product_type=Product.PRODUCT_TYPE_ACCESSORY
        ).prefetch_related('variants').order_by('product_id')

        ids_param = request.query_params.get('ids', '')
        search = request.query_params.get('search', '')
        brand = request.query_params.get('brand', '')
        if ids_param:
            id_list = [i.strip() for i in ids_param.split(',') if i.strip()]
            products = products.filter(id__in=id_list)
        else:
            if search:
                products = products.filter(commercial_name__icontains=search) | \
                           products.filter(brand__icontains=search)
            if brand:
                products = products.filter(brand__iexact=brand)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Accessories Item'

        headers = [
            'Barcode', 'Product name', 'Specification', 'Price', 'Original Price',
            'Image Url', 'Discount Price',
            'Customize1', 'Customize2', 'Customize3', 'Customize4',
            'Customize5', 'Customize6', 'Customize7', 'Customize8',
            'Customize9', 'Customize10', 'Customize11', 'Customize12', 'Customize13',
            'Customize14', 'Customize15', 'Customize16', 'Customize17', 'Customize18',
            'Customize19', 'Customize20',
            'Customize21', 'Customize22', 'Customize23', 'Customize24', 'Customize25',
        ]
        ws.append(headers)

        for p in products:
            variant = p.variants.first()
            srp = fmt_price(variant.unit_price if variant else None)
            image_url = brand_logos.get(p.brand.lower(), '')

            ws.append([
                str(p.product_id),  # Barcode
                p.commercial_name,  # Product name
                p.colour,           # Specification
                srp,                # Price (SRP)
                '',                 # Original Price
                image_url,          # Image Url
                '',                 # Discount Price
                '', '', '', '',     # Customize 1-4
                '', '', '', '',     # Customize 5-8
                '', '', '', '', '', # Customize 9-13
                '', '', '', '', '', # Customize 14-18
                '', '',             # Customize 19-20
                '', '', '', '', '', # Customize 21-25
            ])
            # Force Barcode cell to Text format so numeric-looking IDs don't become scientific notation
            ws.cell(row=ws.max_row, column=1).number_format = '@'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="esl_accessories_export.xlsx"'
        return response


class DebugCSVHeadersView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)
        raw = file.read()
        for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
            try:
                content = raw.decode(enc)
                break
            except Exception:
                continue
        reader = csv.DictReader(io.StringIO(content))
        headers = reader.fieldnames or []
        first_row = next(iter(reader), {})
        return Response({
            'headers': headers,
            'headers_repr': [repr(h) for h in headers],
            'first_row_srp': first_row.get('SRP', 'KEY_NOT_FOUND'),
            'first_row_keys': list(first_row.keys())[:10],
        })
