import io
from openpyxl import load_workbook
from rest_framework import generics, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import Site, LayoutSegment, ESLDevice, ESLTemplate
from .serializers import SiteSerializer, SiteListSerializer, LayoutSegmentSerializer, ESLDeviceSerializer, ESLTemplateSerializer
from products.models import Product

MAINTENANCE_TEMPLATES = {34, 35}


def resolve_template(product):
    """Return auto template_id based on product type and variant count, using ESLTemplate DB."""
    if product.product_type == Product.PRODUCT_TYPE_ACCESSORY:
        t = ESLTemplate.objects.filter(category=ESLTemplate.CATEGORY_ACCESSORY).first()
        return str(t.template_id) if t else ''
    count = product.variants.count()
    t = ESLTemplate.objects.filter(category=ESLTemplate.CATEGORY_DEMO, variant_count=count).first()
    return str(t.template_id) if t else ''


class ESLTemplateListView(generics.ListCreateAPIView):
    queryset = ESLTemplate.objects.all()
    serializer_class = ESLTemplateSerializer
    permission_classes = [IsAuthenticated]


class ESLTemplateDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ESLTemplate.objects.all()
    serializer_class = ESLTemplateSerializer
    permission_classes = [IsAuthenticated]


class SiteListCreateView(generics.ListCreateAPIView):
    queryset = Site.objects.prefetch_related('segments').all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return SiteListSerializer
        return SiteSerializer


class SiteDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Site.objects.prefetch_related('segments').all()
    serializer_class = SiteSerializer
    permission_classes = [IsAuthenticated]


class SegmentListCreateView(generics.ListCreateAPIView):
    serializer_class = LayoutSegmentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return LayoutSegment.objects.filter(site_id=self.kwargs['site_id'])

    def perform_create(self, serializer):
        serializer.save(site_id=self.kwargs['site_id'])


class SegmentDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = LayoutSegmentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return LayoutSegment.objects.filter(site_id=self.kwargs['site_id'])


class ESLDeviceListView(generics.ListAPIView):
    serializer_class = ESLDeviceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return ESLDevice.objects.filter(
            segment_id=self.kwargs['segment_id']
        ).select_related('product')


class ESLDeviceDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ESLDeviceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return ESLDevice.objects.filter(segment_id=self.kwargs['segment_id'])


class ESLDeviceBindProductView(APIView):
    """
    PATCH body options:
      { "product_id": "DEMOITEM001" }  → bind product, auto-set barcode & template
      { "product_id": null }           → unbind, clear barcode & template
      { "template": "34" }             → override template to maintenance, clear barcode
      { "template": "35" }             → override template to maintenance, clear barcode
    """
    permission_classes = [IsAuthenticated]

    def patch(self, request, segment_id, pk):
        try:
            device = ESLDevice.objects.get(pk=pk, segment_id=segment_id)
        except ESLDevice.DoesNotExist:
            return Response({'detail': 'Device not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Override to maintenance template (34 or 35)
        if 'template' in request.data and 'product_id' not in request.data:
            template_val = str(request.data['template'])
            if template_val not in ('34', '35'):
                return Response({'detail': 'Override template hanya boleh 34 atau 35.'}, status=status.HTTP_400_BAD_REQUEST)
            device.template = template_val
            device.barcode = ''
            device.product = None
            device.save()
            return Response(ESLDeviceSerializer(device).data)

        product_id = request.data.get('product_id')

        # Unbind
        if product_id is None:
            device.product = None
            device.barcode = ''
            device.template = ''
            device.save()
            return Response(ESLDeviceSerializer(device).data)

        # Bind product
        try:
            product = Product.objects.prefetch_related('variants').get(product_id=product_id)
        except Product.DoesNotExist:
            return Response({'detail': f'Product {product_id} tidak ditemukan.'}, status=status.HTTP_400_BAD_REQUEST)

        device.product = product
        device.barcode = product.product_id
        device.template = resolve_template(product)
        device.save()
        return Response(ESLDeviceSerializer(device).data)


class ESLDeviceBatchActionView(APIView):
    """
    POST body:
      { "ids": [1,2,3], "action": "unbind" }
      { "ids": [1,2,3], "action": "maintenance", "template": 34 }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, segment_id):
        ids = request.data.get('ids', [])
        action = request.data.get('action')

        if not ids or action not in ('unbind', 'maintenance'):
            return Response({'detail': 'Parameter tidak valid.'}, status=status.HTTP_400_BAD_REQUEST)

        devices = ESLDevice.objects.filter(segment_id=segment_id, id__in=ids)

        if action == 'unbind':
            devices.update(product=None, barcode='', template='')
        else:
            template_val = str(request.data.get('template', ''))
            if template_val not in ('34', '35'):
                return Response({'detail': 'Template maintenance harus 34 atau 35.'}, status=status.HTTP_400_BAD_REQUEST)
            devices.update(product=None, barcode='', template=template_val)

        return Response({'updated': devices.count()})


class ESLDeviceImportView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, segment_id):
        try:
            segment = LayoutSegment.objects.get(pk=segment_id)
        except LayoutSegment.DoesNotExist:
            return Response({'detail': 'Segment not found.'}, status=status.HTTP_404_NOT_FOUND)

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'File wajib diupload.'}, status=status.HTTP_400_BAD_REQUEST)

        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'detail': 'File harus berformat Excel (.xlsx atau .xls).'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
            ws = wb.active
        except Exception:
            return Response({'detail': 'File Excel tidak valid atau rusak.'}, status=status.HTTP_400_BAD_REQUEST)

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({'detail': 'File Excel kosong.'}, status=status.HTTP_400_BAD_REQUEST)

        # Normalize header row
        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        col = {h: i for i, h in enumerate(headers)}

        if 'Code' not in col:
            return Response(
                {'detail': 'Kolom wajib tidak ditemukan: Code'},
                status=status.HTTP_400_BAD_REQUEST
            )

        created, updated, errors = 0, 0, []
        for i, row in enumerate(rows[1:], start=2):
            def get(name):
                idx = col.get(name)
                if idx is None:
                    return ''
                val = row[idx]
                return str(val).strip() if val is not None else ''

            code = get('Code')
            if not code:
                errors.append(f'Baris {i}: Code wajib diisi.')
                continue

            barcode = get('Barcode').lstrip("'")  # strip leading apostrophe if present
            template_val = get('Templates')

            # Try to bind product by product_id matching barcode
            product = None
            if barcode:
                product = Product.objects.filter(product_id=barcode).first()
                # If product found and no template specified, auto-resolve template
                if product and not template_val:
                    template_val = resolve_template(product)

            device, was_created = ESLDevice.objects.update_or_create(
                segment=segment,
                code=code,
                defaults={
                    'barcode': barcode,
                    'template': template_val,
                    'ap': get('Ap'),
                    'desc': get('Desc'),
                    'product': product,
                }
            )
            if was_created:
                created += 1
            else:
                updated += 1

        return Response({
            'created': created,
            'updated': updated,
            'errors': errors,
        }, status=status.HTTP_200_OK)
